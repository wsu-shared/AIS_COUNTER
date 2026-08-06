"""Reporting: an annotated PNG next to a multi-sheet XLSX.

The PNG shows the raw image with every accepted trace drawn over it and labelled with its
length, so a reviewer can check the numbers against the pixels without opening MATLAB. The
XLSX carries the per-AIS table, a per-image summary, and the exact parameters used.

Excluded AIS are written to the workbook too, with the reason, so a rejected measurement is
always auditable rather than silently absent.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import numpy as np

from .measure import PIXCONV, circularity

# Drawn in matplotlib, which is imported lazily in render_png: it is the slowest import in the
# project and nothing else here needs it, so a --batch --no-png run never pays for it.
#
# The accepted-trace colour is config.skeleton_color; these two stay fixed so that "rejected"
# and "landmark" keep meaning the same thing whatever the user picks for the trace itself.
EXCLUDED_COLOUR = "#FF5252"
LABEL_COLOUR = "#FFEB3B"

ROW_HEADERS = [
    "image", "ais", "uid", "included", "source", "length_um", "length_mode",
    "ais_length_um", "trace_length_um", "arclength_um",
    "circularity", "start_um", "end_um", "mid_um", "max_um", "profile_points",
    "seed_row", "seed_col", "component_label", "threshold", "note",
]
"""Column order for the per-AIS table, shared by the workbook and the CSV so the two
never drift apart.

``length_um`` is whichever measurement was chosen as the headline and ``length_mode`` says
which -- a column of lengths without it is not comparable to anything. The three that follow
are the alternatives, always present whatever the mode, so a table measured one way can still
answer a question asked the other way. ``start_um``/``end_um``/``mid_um``/``max_um`` are the
original's own five (with ``ais_length_um``), and never change with the mode."""

PNG_LINEWIDTH_SCALE = 0.8
"""Screen pixels -> matplotlib points for the drawn trace.

Chosen so the default width of 2.0 px reproduces the 1.6 pt line every report has used so
far; changing the slider then scales from there."""


@dataclass
class ReportPaths:
    png: Path
    xlsx: Path


def _autoscale(img: np.ndarray, low: float = 0.5, high: float = 99.5):
    """Percentile contrast stretch -- raw 16-bit AIS frames are unreadable unscaled."""
    a = np.asarray(img, dtype=np.float64)
    lo, hi = np.percentile(a, [low, high])
    if hi <= lo:
        lo, hi = a.min(), a.max()
    if hi <= lo:
        return np.zeros_like(a)
    return np.clip((a - lo) / (hi - lo), 0, 1)


def render_png(result, path, dpi: int = 150, show_excluded: bool = False) -> Path:
    """Draw every accepted trace over the raw image, labelled with its length."""
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    path = Path(path)
    img = result.image.raw
    h, w = img.shape

    fig_w = 14.0
    fig_h = fig_w * h / w
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.imshow(_autoscale(img), cmap="gray", interpolation="nearest")
    ax.set_xlim(0, w)
    ax.set_ylim(h, 0)
    ax.axis("off")

    trace_colour = getattr(result.config, "skeleton_color", None) or "magenta"
    width = float(getattr(result.config, "skeleton_width", None) or 2.0)
    linewidth = max(0.2, width * PNG_LINEWIDTH_SCALE)
    records = result.records if show_excluded else result.active
    for record in records:
        m = record.measurement
        colour = EXCLUDED_COLOUR if record.excluded else trace_colour
        ax.plot(m.spline_x, m.spline_y, "-", color=colour, linewidth=linewidth, alpha=0.9)

        # The two ends of what the length actually measures: the f crossings in the original's
        # mode, the ends of the trace in the whole-trace modes (see measure.LENGTH_MODES).
        if not record.excluded:
            for idx, marker in ((m.ais_start_idx, "o"), (m.ais_end_idx, "s")):
                i = int(np.clip(idx - 1, 0, m.x_pix.size - 1))
                ax.plot(
                    m.x_pix[i], m.y_pix[i], marker, color=LABEL_COLOUR,
                    markersize=4, markeredgewidth=0,
                )

        label_x = float(m.x_pix[0])
        label_y = float(m.y_pix[0])
        text = f"{record.index}: {m.length_um:.1f}um"
        if record.excluded:
            text = f"excluded ({m.length_um:.1f}um)"
        ax.annotate(
            text,
            xy=(label_x, label_y),
            xytext=(6, -6),
            textcoords="offset points",
            color=LABEL_COLOUR if not record.excluded else EXCLUDED_COLOUR,
            fontsize=7,
            fontweight="bold",
            path_effects=_stroke(),
        )

    lengths = result.lengths
    if lengths.size:
        summary = f"{result.image.name}   n={lengths.size}   mean={lengths.mean():.2f} um"
    else:
        summary = f"{result.image.name}   n=0"
    # Only when it is not the original's: a PNG travels on its own into slides and emails, and
    # a number that is not the AIS length has to say so or nobody can reconcile it with a
    # previous figure. Silent in "profile" mode, which is what every number has always meant.
    labels = {
        "trace": "[reported: whole trace, pixel steps]",
        "arclength": "[reported: whole trace, spline arc length]",
        "max": "[reported: AIS max position, not a length]",
    }
    mode = getattr(result.config, "length_mode", "profile")
    if mode in labels:
        summary += "   " + labels[mode]
    if result.image.processed_is_derived:
        summary += "   [processed channel DERIVED - not ImageJ 'method 2.5']"
    ax.set_title(summary, fontsize=10, color="black", pad=8)

    fig.tight_layout()
    fig.savefig(path, dpi=dpi, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


def _stroke():
    import matplotlib.patheffects as pe

    return [pe.withStroke(linewidth=2.0, foreground="black")]


def _rows_for(result) -> list:
    rows = []
    for record in result.records:
        m = record.measurement
        rows.append(
            {
                "image": result.image.name,
                "ais": record.index if not record.excluded else None,
                "uid": record.uid,
                "included": not record.excluded,
                "source": record.source,
                "length_um": round(m.length_um, 4),
                "length_mode": m.length_mode,
                "ais_length_um": round(m.ais_length_um, 4),
                "trace_length_um": round(m.trace_length_um, 4),
                "arclength_um": round(m.arclength_um, 4),
                "circularity": round(circularity(m), 4),
                "start_um": round(m.start_um, 4),
                "end_um": round(m.end_um, 4),
                "mid_um": round(m.mid_um, 4),
                "max_um": round(m.max_um, 4),
                "profile_points": m.n_profile_points,
                "seed_row": int(m.seed_rc[0]),
                "seed_col": int(m.seed_rc[1]),
                # A joined AIS spans several components; listing all of them is what makes a
                # merge traceable back to the segmentation afterwards.
                "component_label": (
                    "+".join(str(v) for v in sorted(record.labels))
                    if len(record.labels) > 1
                    else record.label
                ),
                # The threshold this AIS was measured at, which is the image's own except
                # under rethreshold="original", where every AIS has its own. Falls back to
                # the image threshold for joined and spliced traces, which do not record one.
                "threshold": round(
                    float(record.threshold or result.segmentation.threshold), 6
                ),
                "note": record.reason or "; ".join(m.warnings),
            }
        )
    return rows


def write_xlsx(results, path, config=None) -> Path:
    """Write the workbook: per-AIS rows, a per-image summary, and the parameters used.

    *results* may be a single ``AnalysisResult`` or a sequence of them.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not isinstance(results, (list, tuple)):
        results = [results]
    path = Path(path)

    wb = Workbook()

    # --- sheet 1: every AIS -------------------------------------------------------
    ws = wb.active
    ws.title = "AIS"
    ws.append(ROW_HEADERS)
    headers = ROW_HEADERS
    excluded_fill = PatternFill("solid", fgColor="FFE0E0")
    for result in results:
        for row in _rows_for(result):
            ws.append([row[h] for h in headers])
            if not row["included"]:
                for cell in ws[ws.max_row]:
                    cell.fill = excluded_fill

    # --- sheet 2: one row per image ----------------------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2.append(
        ["image", "n_ais", "mean_length_um", "median_length_um", "sd_length_um",
         "min_length_um", "max_length_um", "n_excluded", "threshold", "rethreshold",
         "length_mode", "pixconv_um", "processed_derived", "image_path"]
    )
    for result in results:
        L = result.lengths
        n_excl = sum(1 for r in result.records if r.excluded)
        ws2.append([
            result.image.name,
            int(L.size),
            round(float(L.mean()), 4) if L.size else None,
            round(float(np.median(L)), 4) if L.size else None,
            round(float(L.std(ddof=1)), 4) if L.size > 1 else None,
            round(float(L.min()), 4) if L.size else None,
            round(float(L.max()), 4) if L.size else None,
            n_excl,
            round(float(result.segmentation.threshold), 6),
            # In "original" mode the image-level threshold above is only the starting point:
            # each AIS was measured at its own, in the per-AIS sheet's threshold column.
            result.rethreshold,
            result.config.length_mode,
            result.config.pixconv or result.image.pixconv,
            result.image.processed_is_derived,
            # Which file on disk this row came from. Image names repeat between folders, and a
            # summary that cannot be traced back to a file is hard to trust six months later.
            str(Path(result.image.path).resolve()),
        ])

    # --- sheet 3: provenance ------------------------------------------------------
    ws3 = wb.create_sheet("Parameters")
    cfg = config or (results[0].config if results else None)
    ws3.append(["parameter", "value"])
    ws3.append(["generated", datetime.now().isoformat(timespec="seconds")])
    ws3.append(["aiscounter_version", _version()])
    ws3.append(["source", "port of original/ais_auto.m, validated against MATLAB R2024b"])
    if cfg is not None:
        for key, value in vars(cfg).items():
            ws3.append([key, str(value)])
    ws3.append(["pixconv_default_um", PIXCONV])

    for sheet in (ws, ws2, ws3):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left")
        sheet.freeze_panes = "A2"
        for i, column in enumerate(sheet.columns, start=1):
            width = max((len(str(c.value)) for c in column if c.value is not None), default=8)
            sheet.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_csv(results, path) -> Path:
    """Write the per-AIS table as CSV. *results* is one ``AnalysisResult`` or a sequence.

    The same rows and column order as the workbook's first sheet, in the format anything
    reads without openpyxl. Written to a temporary file and moved into place, because this
    is what autosave calls on a timer: a crash or a Ctrl-C mid-write must not be able to
    leave a half-written file where a complete one used to be, and ``os.replace`` is atomic
    within a filesystem.
    """
    import csv
    import os
    import tempfile

    if not isinstance(results, (list, tuple)):
        results = [results]
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    handle, tmp = tempfile.mkstemp(dir=str(path.parent), prefix=f".{path.name}.", suffix=".tmp")
    try:
        # mkstemp creates 0600. These sit in a shared results folder next to the workbooks,
        # which are written with ordinary permissions, so a CSV nobody else on the machine
        # can open would be a surprise nothing in the UI explains.
        umask = os.umask(0)
        os.umask(umask)
        os.chmod(tmp, 0o666 & ~umask)

        with os.fdopen(handle, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=ROW_HEADERS, extrasaction="ignore")
            writer.writeheader()
            for result in results:
                writer.writerows(_rows_for(result))
        os.replace(tmp, path)
    except BaseException:
        Path(tmp).unlink(missing_ok=True)
        raise
    return path


def _version() -> str:
    from . import __version__

    return __version__


def write_report(result, outdir=None, stem: str | None = None) -> ReportPaths:
    """Write the PNG + XLSX pair for a single image, named after it."""
    outdir = Path(outdir) if outdir else Path(result.image.path).parent
    outdir.mkdir(parents=True, exist_ok=True)
    stem = stem or result.image.name

    png = render_png(result, outdir / f"{stem}_ais_traces.png")
    xlsx = write_xlsx(result, outdir / f"{stem}_ais_results.xlsx", config=result.config)
    return ReportPaths(png=png, xlsx=xlsx)
